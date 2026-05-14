from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .base import BaseParser, PageContent, ParsedDocument, extract_norm_codes

_SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}


class XlsxParser(BaseParser):
    """Parser for Excel files (.xlsx, .xlsm) using openpyxl.

    Each worksheet becomes one PageContent. Rows are serialized as
    pipe-separated cells so the text is human-readable and searchable.
    """

    @classmethod
    def supports(cls, path: Path) -> bool:
        return path.suffix.lower() in _SUPPORTED_EXTENSIONS

    def parse(self, path: Path) -> ParsedDocument:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        pages: list[PageContent] = []

        try:
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                lines: list[str] = []

                for row in ws.iter_rows(values_only=True):
                    cells = [
                        str(cell).strip() for cell in row if cell is not None and str(cell).strip()
                    ]
                    if cells:
                        lines.append(" | ".join(cells))

                if lines:
                    pages.append(
                        PageContent(
                            page_number=sheet_idx,
                            text=f"[Hoja: {sheet_name}]\n" + "\n".join(lines),
                        )
                    )
        finally:
            wb.close()

        full_text = "\n\n".join(p.text for p in pages)
        norm_codes = sorted(set(extract_norm_codes(full_text)) | set(extract_norm_codes(path.stem)))

        return ParsedDocument(
            source_path=path,
            doc_type="xlsx",
            title=path.stem,
            pages=pages,
            norm_codes=norm_codes,
        )
